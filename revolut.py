"""
Revolut import — parses a raw Revolut account-statement export
(CSV or XLSX) and pre-classifies each transaction into Fluxo's
category tree with merchant-keyword rules.

No LLM: the merchant set is small and repetitive, so deterministic
rules cover the bulk; the preview screen handles the tail by hand.
"""
import io
import csv
from datetime import datetime, date

# (keyword in description, (cat1, cat2, cat3), restrict to tx_type or None)
# First match wins — order matters.
RULES = [
    ('relai',              ('Investments', 'Bitcoin DCA', ''), None),
    ('sammelkartenmarkt',  ('Investments', 'Business', ''), None),
    ('cardmarket',         ('Investments', 'Business', ''), None),
    ('mangopay',           ('Income', 'Business', 'Vinted'), 'inflow'),
    ('vinted',             ('Investments', 'Business', ''), 'outflow'),
    ('do invest',          ('Food', 'Snack', ''), None),
    ('auchan',             ('Food', 'Groceries', ''), None),
    ('delhaize',           ('Food', 'Groceries', ''), None),
    ('carrefour',          ('Food', 'Groceries', ''), None),
    ('mcdonald',           ('Food', 'Dinner', ''), None),
    ('flex',               ('Transportation', 'Car Rental', ''), None),
    ('uber',               ('Transportation', 'Uber', ''), None),
    ("vel'oh",             ('Transportation', 'Bike', ''), None),
    ('zara',               ('Miscellaneous', 'Shopping', 'Clothes'), None),
    ('uniqlo',             ('Miscellaneous', 'Shopping', 'Clothes'), None),
    ('nike',               ('Miscellaneous', 'Shopping', 'Clothes'), None),
    ('h&m',                ('Miscellaneous', 'Shopping', 'Clothes'), None),
    ('adventure',          ('Miscellaneous', 'Shopping', 'Clothes'), None),
    ('dhgate',             ('Miscellaneous', 'Shopping', 'Tech'), None),
    ('aliexpress',         ('Miscellaneous', 'Shopping', 'Tech'), None),
    ('amazon',             ('Miscellaneous', 'Shopping', 'Tech'), None),
    ('barber',             ('Miscellaneous', 'Personal', 'Barbershop'), None),
    ('utopia',             ('Miscellaneous', 'Entertainment', 'Tickets'), None),
    ('kiosk',              ('Food', 'Snacks', ''), None),
    ('1global',            ('Subscriptions', 'Phone Plan', ''), None),
    ('revolut bank',       ('Investments', 'Savings', ''), None),
    ('transfer from',      ('Miscellaneous', 'Gifts', ''), 'inflow'),
    ('transfer to',        ('Miscellaneous', 'Gifts', ''), 'outflow'),
]

DEFAULT_OUTFLOW = ('Miscellaneous', 'Unknown', '')
DEFAULT_INFLOW = ('Income', '', '')

# Internal money movements — not real income/spend; pre-deselected in preview.
INTERNAL_PATTERNS = ('apple pay deposit', 'payment from iandolo')


def classify(description, tx_type):
    d = (description or '').lower()
    for key, path, only in RULES:
        if key in d and (only is None or only == tx_type):
            return path
    return DEFAULT_INFLOW if tx_type == 'inflow' else DEFAULT_OUTFLOW


def _cell(v):
    """Normalize a cell value to a trimmed string."""
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    return str(v).strip()


def _read_rows(file_bytes, filename):
    """Return a list of row-lists from a Revolut CSV or XLSX export."""
    name = (filename or '').lower()
    if name.endswith('.xlsx'):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        return [[_cell(c) for c in row] for row in ws.iter_rows(values_only=True)]
    text = file_bytes.decode('utf-8-sig', errors='replace')
    return [[_cell(c) for c in row] for row in csv.reader(io.StringIO(text))]


def parse_export(file_bytes, filename):
    """Parse a raw Revolut export into preview rows.

    Returns a list of dicts:
      date, merchant, transaction_type, amount, category_1/2/3,
      include (bool), note (why pre-excluded, if any)
    """
    rows = _read_rows(file_bytes, filename)
    if not rows:
        return []

    header = [h.lower() for h in rows[0]]

    def idx(name):
        return header.index(name) if name in header else None

    i_completed = idx('completed date')
    i_desc = idx('description')
    i_amount = idx('amount')
    i_state = idx('state')
    if i_desc is None or i_amount is None:
        raise ValueError('Not a Revolut export — missing Description/Amount columns')

    out = []
    for r in rows[1:]:
        if not any(x for x in r):
            continue

        def g(i):
            return r[i] if i is not None and i < len(r) else ''

        state = g(i_state).upper()
        completed = g(i_completed)
        try:
            amount = float(g(i_amount).replace(',', '.'))
        except ValueError:
            continue
        merchant = g(i_desc)
        tx_type = 'inflow' if amount > 0 else 'outflow'

        include, note = True, ''
        if state and state != 'COMPLETED':
            include, note = False, state.lower()  # pending / reverted
        elif any(p in merchant.lower() for p in INTERNAL_PATTERNS):
            include, note = False, 'internal top-up'

        c1, c2, c3 = classify(merchant, tx_type)
        out.append({
            'date': completed[:10] if completed else '',
            'merchant': merchant,
            'transaction_type': tx_type,
            'amount': round(abs(amount), 2),
            'category_1': c1, 'category_2': c2, 'category_3': c3,
            'include': include,
            'note': note,
        })
    return out
